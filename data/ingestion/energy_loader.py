from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


CITY_POPULATION_WEIGHTS = {
    "Delhi": 33800000,
    "Mumbai": 21300000,
    "Kolkata": 15500000,
    "Bengaluru": 13600000,
    "Chennai": 12300000,
    "Hyderabad": 11100000,
    "Ahmedabad": 8800000,
    "Pune": 7400000,
    "Surat": 7200000,
    "Jaipur": 4100000,
    "Lucknow": 3900000,
    "Kanpur": 3200000,
    "Nagpur": 3000000,
    "Indore": 2700000,
    "Patna": 2600000,
    "Bhopal": 2400000,
    "Visakhapatnam": 2300000,
    "Ludhiana": 2200000,
    "Vadodara": 2200000,
    "Coimbatore": 2200000,
    "Kochi": 2100000,
    "Rajkot": 1800000,
    "Thiruvananthapuram": 1700000,
    "Vijayawada": 1600000,
    "Srinagar": 1500000,
    "Amritsar": 1400000,
    "Bhubaneswar": 1200000,
    "Guwahati": 1200000,
    "Chandigarh": 1200000,
    "Jammu": 750000,
}


def convert_energy_xlsx_folder(source_dir: str | Path, output_path: str | Path, cities_path: str | Path) -> Path:
    source_dir = Path(source_dir)
    output_path = Path(output_path)
    cities_path = Path(cities_path)
    xlsx_files = sorted(source_dir.glob("*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(f"No .xlsx files found in {source_dir}")

    hourly_frames = [_extract_hourly_national_demand(path) for path in xlsx_files]
    national = pd.concat(hourly_frames, ignore_index=True)
    national = national.groupby("observed_at", as_index=False)["total_demand_mwh"].mean()
    energy = _allocate_to_cities(national, cities_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    energy.to_csv(output_path, index=False)
    return output_path


def _extract_hourly_national_demand(path: Path) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        labels = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        demand_index = _find_demand_column(header, labels)
        rows = []
        for row in sheet.iter_rows(min_row=3, values_only=True):
            timestamp = row[0]
            demand = row[demand_index] if demand_index < len(row) else None
            if timestamp is not None and demand is not None:
                rows.append((timestamp, demand))
    finally:
        workbook.close()

    frame = pd.DataFrame(rows, columns=["observed_at", "total_demand_mwh"])
    frame["observed_at"] = pd.to_datetime(frame["observed_at"], errors="coerce", utc=True)
    frame["total_demand_mwh"] = pd.to_numeric(frame["total_demand_mwh"], errors="coerce")
    frame = frame.dropna()
    return frame.set_index("observed_at").resample("1h").mean().reset_index()


def _find_demand_column(header: tuple, labels: tuple) -> int:
    for index, label in enumerate(labels):
        if str(label).strip() == "NLDC_DEMAND|P":
            return index
    for index, col in enumerate(header):
        if str(col).strip() == "SCADA/ANALOG/044MQ067/0":
            return index
    raise ValueError("Could not find NLDC demand column in workbook")


def _allocate_to_cities(national: pd.DataFrame, cities_path: Path) -> pd.DataFrame:
    cities = pd.read_csv(cities_path)
    cities["population_weight"] = cities["city"].map(CITY_POPULATION_WEIGHTS).fillna(1000000)
    cities["weight"] = cities["population_weight"] / cities["population_weight"].sum()
    frames = []
    for _, city in cities.iterrows():
        city_frame = national.copy()
        city_frame["city"] = city["city"]
        city_frame["energy_demand_mwh"] = city_frame["total_demand_mwh"] * city["weight"]
        frames.append(city_frame[["observed_at", "city", "energy_demand_mwh"]])
    return pd.concat(frames, ignore_index=True)
